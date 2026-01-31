"""
Tests for M1a - CSV and Excel Session Exporters.

Tests cover:
- CSV: format props, export basic, contains session_id/metrics, sections, exclude flags, empty session
- Excel: format props, returns bytes, valid xlsx (openpyxl.load_workbook), sheets present, exclude flags
- Factory: get_exporter("csv"), get_exporter("xlsx"), get_exporter("excel")
"""

import csv
import io
import pytest
from datetime import datetime, timezone

from app.services.session_exporters import get_exporter
from app.services.session_exporters.base import (
    BaseSessionExporter,
    SessionData,
    ThinkingBlockExport,
    ToolExecutionExport,
    MessageExport,
)
from app.services.session_exporters.csv_exporter import CSVSessionExporter
from app.services.session_exporters.excel_exporter import ExcelSessionExporter


# =============================================================================
# FIXTURES
# =============================================================================


def _make_session_data(
    thinking=True, tools=True, messages=True
) -> SessionData:
    """Create sample session data for testing."""
    now = datetime(2024, 1, 15, 10, 30, 0, tzinfo=timezone.utc)

    thinking_blocks = []
    if thinking:
        thinking_blocks = [
            ThinkingBlockExport(
                block_type="thinking",
                content="Analyzing the user's request...",
                token_count=150,
                sequence_number=1,
                signature="sig-001",
                content_hash="hash-001",
            ),
            ThinkingBlockExport(
                block_type="reflection",
                content="Let me reconsider...",
                token_count=80,
                sequence_number=2,
            ),
        ]

    tool_executions = []
    if tools:
        tool_executions = [
            ToolExecutionExport(
                tool_name="read_file",
                tool_input={"path": "/src/main.py"},
                tool_output="def main(): pass",
                duration_ms=120,
                success=True,
                executed_at=now,
            ),
            ToolExecutionExport(
                tool_name="write_file",
                tool_input={"path": "/src/test.py", "content": "import pytest"},
                tool_output=None,
                duration_ms=50,
                success=False,
                error_message="Permission denied",
                executed_at=now,
            ),
        ]

    msg_list = []
    if messages:
        msg_list = [
            MessageExport(
                role="user",
                content="Hello, can you help me?",
                message_id="msg-001",
                timestamp=now,
            ),
            MessageExport(
                role="assistant",
                content="Of course! What do you need?",
                message_id="msg-002",
                parent_id="msg-001",
                timestamp=now,
            ),
        ]

    return SessionData(
        session_id="test-session-123",
        created_at=now,
        provider="anthropic",
        model="claude-3-opus",
        thinking_blocks=thinking_blocks,
        tool_executions=tool_executions,
        messages=msg_list,
        total_input_tokens=1000,
        total_output_tokens=500,
        cache_creation_tokens=200,
        cache_read_tokens=300,
        total_duration_ms=5000,
        total_cost=0.05,
        metadata={"version": "1.0"},
    )


def _make_empty_session() -> SessionData:
    """Create empty session data."""
    return SessionData(
        session_id="empty-session",
        created_at=datetime(2024, 1, 1, 0, 0, 0, tzinfo=timezone.utc),
    )


# =============================================================================
# CSV EXPORTER - FORMAT PROPERTIES
# =============================================================================


class TestCSVExporterProperties:
    """Tests for CSV exporter format properties."""

    def test_format_name(self):
        """Should return 'csv'."""
        exporter = CSVSessionExporter()
        assert exporter.format_name == "csv"

    def test_content_type(self):
        """Should return text/csv."""
        exporter = CSVSessionExporter()
        assert exporter.content_type == "text/csv"

    def test_file_extension(self):
        """Should return 'csv'."""
        exporter = CSVSessionExporter()
        assert exporter.file_extension == "csv"

    def test_is_base_session_exporter(self):
        """Should be a BaseSessionExporter subclass."""
        exporter = CSVSessionExporter()
        assert isinstance(exporter, BaseSessionExporter)

    def test_get_filename(self):
        """Should generate a filename with .csv extension."""
        exporter = CSVSessionExporter()
        filename = exporter.get_filename("test-123")
        assert filename.endswith(".csv")
        assert "test-123" in filename


# =============================================================================
# CSV EXPORTER - EXPORT
# =============================================================================


class TestCSVExporterExport:
    """Tests for CSV export functionality."""

    def test_export_basic(self):
        """Should export valid CSV string."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert isinstance(result, str)
        assert len(result) > 0

    def test_export_contains_session_id(self):
        """Should contain the session ID."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "test-session-123" in result

    def test_export_contains_metrics(self):
        """Should contain metrics values."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "1000" in result  # input tokens
        assert "500" in result  # output tokens
        assert "5000" in result  # duration

    def test_export_has_metadata_section(self):
        """Should have metadata section header."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "SESSION METADATA" in result

    def test_export_has_metrics_section(self):
        """Should have metrics section header."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "METRICS" in result

    def test_export_has_thinking_section(self):
        """Should have thinking blocks section."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "THINKING BLOCKS" in result
        assert "Analyzing the user" in result

    def test_export_has_tools_section(self):
        """Should have tool executions section."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "TOOL EXECUTIONS" in result
        assert "read_file" in result

    def test_export_has_messages_section(self):
        """Should have messages section."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert "MESSAGES" in result
        assert "Hello, can you help me?" in result

    def test_export_exclude_thinking(self):
        """Should exclude thinking blocks when flag is False."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_thinking=False)

        assert "THINKING BLOCKS" not in result
        assert "Analyzing the user" not in result

    def test_export_exclude_tools(self):
        """Should exclude tool executions when flag is False."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_tools=False)

        assert "TOOL EXECUTIONS" not in result

    def test_export_exclude_messages(self):
        """Should exclude messages when flag is False."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_messages=False)

        assert "MESSAGES" not in result

    def test_export_empty_session(self):
        """Should handle empty session gracefully."""
        exporter = CSVSessionExporter()
        session = _make_empty_session()
        result = exporter.export(session)

        assert isinstance(result, str)
        assert "empty-session" in result
        # No thinking/tools/messages sections for empty data
        assert "THINKING BLOCKS" not in result
        assert "TOOL EXECUTIONS" not in result

    def test_export_is_valid_csv(self):
        """Output should be parseable as CSV."""
        exporter = CSVSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        assert len(rows) > 5  # At least metadata + metrics


# =============================================================================
# EXCEL EXPORTER - FORMAT PROPERTIES
# =============================================================================


class TestExcelExporterProperties:
    """Tests for Excel exporter format properties."""

    def test_format_name(self):
        """Should return 'xlsx'."""
        exporter = ExcelSessionExporter()
        assert exporter.format_name == "xlsx"

    def test_content_type(self):
        """Should return correct MIME type."""
        exporter = ExcelSessionExporter()
        assert exporter.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_file_extension(self):
        """Should return 'xlsx'."""
        exporter = ExcelSessionExporter()
        assert exporter.file_extension == "xlsx"

    def test_is_base_session_exporter(self):
        """Should be a BaseSessionExporter subclass."""
        exporter = ExcelSessionExporter()
        assert isinstance(exporter, BaseSessionExporter)

    def test_get_filename(self):
        """Should generate a filename with .xlsx extension."""
        exporter = ExcelSessionExporter()
        filename = exporter.get_filename("test-456")
        assert filename.endswith(".xlsx")
        assert "test-456" in filename


# =============================================================================
# EXCEL EXPORTER - EXPORT
# =============================================================================


class TestExcelExporterExport:
    """Tests for Excel export functionality."""

    def test_export_returns_bytes(self):
        """Should return bytes (not string)."""
        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_valid_xlsx(self):
        """Should produce a valid xlsx file loadable by openpyxl."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) >= 1
        wb.close()

    def test_export_has_overview_sheet(self):
        """Should have an Overview sheet."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        assert "Overview" in wb.sheetnames
        wb.close()

    def test_export_has_thinking_sheet(self):
        """Should have a Thinking Blocks sheet."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        assert "Thinking Blocks" in wb.sheetnames
        wb.close()

    def test_export_has_tools_sheet(self):
        """Should have a Tool Executions sheet."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        assert "Tool Executions" in wb.sheetnames
        wb.close()

    def test_export_has_messages_sheet(self):
        """Should have a Messages sheet."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        assert "Messages" in wb.sheetnames
        wb.close()

    def test_export_overview_contains_session_id(self):
        """Overview sheet should contain the session ID."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session)

        wb = load_workbook(io.BytesIO(result))
        ws = wb["Overview"]
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "test-session-123" in values
        wb.close()

    def test_export_exclude_thinking(self):
        """Should not create Thinking Blocks sheet when excluded."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_thinking=False)

        wb = load_workbook(io.BytesIO(result))
        assert "Thinking Blocks" not in wb.sheetnames
        wb.close()

    def test_export_exclude_tools(self):
        """Should not create Tool Executions sheet when excluded."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_tools=False)

        wb = load_workbook(io.BytesIO(result))
        assert "Tool Executions" not in wb.sheetnames
        wb.close()

    def test_export_exclude_messages(self):
        """Should not create Messages sheet when excluded."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_session_data()
        result = exporter.export(session, include_messages=False)

        wb = load_workbook(io.BytesIO(result))
        assert "Messages" not in wb.sheetnames
        wb.close()

    def test_export_empty_session(self):
        """Should handle empty session gracefully."""
        from openpyxl import load_workbook

        exporter = ExcelSessionExporter()
        session = _make_empty_session()
        result = exporter.export(session)

        assert isinstance(result, bytes)
        wb = load_workbook(io.BytesIO(result))
        assert "Overview" in wb.sheetnames
        wb.close()


# =============================================================================
# FACTORY FUNCTION
# =============================================================================


class TestExporterFactory:
    """Tests for get_exporter factory function."""

    def test_get_csv_exporter(self):
        """get_exporter('csv') should return CSVSessionExporter."""
        exporter = get_exporter("csv")
        assert isinstance(exporter, CSVSessionExporter)

    def test_get_xlsx_exporter(self):
        """get_exporter('xlsx') should return ExcelSessionExporter."""
        exporter = get_exporter("xlsx")
        assert isinstance(exporter, ExcelSessionExporter)

    def test_get_excel_exporter(self):
        """get_exporter('excel') should return ExcelSessionExporter."""
        exporter = get_exporter("excel")
        assert isinstance(exporter, ExcelSessionExporter)

    def test_get_csv_case_insensitive(self):
        """Should be case-insensitive."""
        exporter = get_exporter("CSV")
        assert isinstance(exporter, CSVSessionExporter)

    def test_get_xlsx_case_insensitive(self):
        """Should be case-insensitive."""
        exporter = get_exporter("XLSX")
        assert isinstance(exporter, ExcelSessionExporter)

    def test_invalid_format_raises(self):
        """Should raise ValueError for unsupported format."""
        with pytest.raises(ValueError, match="Unsupported format"):
            get_exporter("pdf")

    def test_existing_formats_still_work(self):
        """Existing formats (json, md, xml, jsonl) should still work."""
        assert get_exporter("json") is not None
        assert get_exporter("md") is not None
        assert get_exporter("xml") is not None
        assert get_exporter("jsonl") is not None
