"""
Excel Session Exporter - Multi-sheet workbook for analysis

Week 61+: Enhanced Observability Integration - Excel Export
"""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.session_exporters.base import BaseSessionExporter, SessionData


class ExcelSessionExporter(BaseSessionExporter):
    """
    Export CCTrace session to Excel (.xlsx) format.

    Produces multi-sheet workbook with:
    - Overview: Session metadata and metrics
    - Thinking Blocks: All thinking block data
    - Tool Executions: All tool execution data
    - Messages: Message history
    """

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return "xlsx"

    def export(
        self,
        session_data: SessionData,
        include_thinking: bool = True,
        include_tools: bool = True,
        include_messages: bool = True,
    ) -> bytes:
        """Export session to Excel format. Returns bytes."""
        wb = Workbook()

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Sheet 1: Overview
        ws_overview = wb.active
        ws_overview.title = "Overview"
        self._write_overview(ws_overview, session_data, header_font, header_fill)

        # Sheet 2: Thinking Blocks
        if include_thinking:
            ws_thinking = wb.create_sheet("Thinking Blocks")
            self._write_thinking_blocks(ws_thinking, session_data, header_font, header_fill)

        # Sheet 3: Tool Executions
        if include_tools:
            ws_tools = wb.create_sheet("Tool Executions")
            self._write_tool_executions(ws_tools, session_data, header_font, header_fill)

        # Sheet 4: Messages
        if include_messages:
            ws_messages = wb.create_sheet("Messages")
            self._write_messages(ws_messages, session_data, header_font, header_fill)

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _apply_header_style(self, ws, row_num, header_font, header_fill):
        """Apply styling to header row."""
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill

    def _write_overview(self, ws, session_data: SessionData, header_font, header_fill):
        """Write overview sheet with metadata and metrics."""
        # Metadata section
        ws.append(["Session Metadata"])
        ws["A1"].font = Font(bold=True, size=14)

        ws.append(["Field", "Value"])
        self._apply_header_style(ws, 2, header_font, header_fill)

        ws.append(["Session ID", session_data.session_id])
        ws.append(["Created At", session_data.created_at.isoformat() if session_data.created_at else ""])
        ws.append(["Provider", session_data.provider])
        ws.append(["Model", session_data.model])
        ws.append(["Export Timestamp", datetime.now(timezone.utc).isoformat()])
        ws.append([])

        # Metrics section
        ws.append(["Metrics"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=14)

        ws.append(["Metric", "Value"])
        self._apply_header_style(ws, ws.max_row, header_font, header_fill)

        ws.append(["Total Input Tokens", session_data.total_input_tokens])
        ws.append(["Total Output Tokens", session_data.total_output_tokens])
        ws.append(["Total Tokens", session_data.total_input_tokens + session_data.total_output_tokens])
        ws.append(["Cache Creation Tokens", session_data.cache_creation_tokens])
        ws.append(["Cache Read Tokens", session_data.cache_read_tokens])
        ws.append(["Total Duration (ms)", session_data.total_duration_ms])
        ws.append(["Total Cost (USD)", session_data.total_cost])

        # Summary counts
        ws.append([])
        ws.append(["Content Summary"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=14)
        ws.append(["Thinking Blocks", len(session_data.thinking_blocks)])
        ws.append(["Tool Executions", len(session_data.tool_executions)])
        ws.append(["Messages", len(session_data.messages)])

        # Auto-width for columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    def _write_thinking_blocks(self, ws, session_data: SessionData, header_font, header_fill):
        """Write thinking blocks sheet."""
        headers = ["Sequence", "Type", "Content", "Tokens", "Signature", "Hash"]
        ws.append(headers)
        self._apply_header_style(ws, 1, header_font, header_fill)

        for block in session_data.thinking_blocks:
            ws.append([
                block.sequence_number,
                block.block_type,
                block.content[:32767],  # Excel cell limit
                block.token_count,
                block.signature or "",
                block.content_hash or "",
            ])

        # Auto-width
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

    def _write_tool_executions(self, ws, session_data: SessionData, header_font, header_fill):
        """Write tool executions sheet."""
        headers = ["Tool Name", "Success", "Duration (ms)", "Error", "Executed At", "Input", "Output"]
        ws.append(headers)
        self._apply_header_style(ws, 1, header_font, header_fill)

        for tool in session_data.tool_executions:
            ws.append([
                tool.tool_name,
                "Yes" if tool.success else "No",
                tool.duration_ms or "",
                tool.error_message or "",
                tool.executed_at.isoformat() if tool.executed_at else "",
                str(tool.tool_input)[:32767],
                (tool.tool_output or "")[:32767],
            ])

        # Auto-width
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 50

    def _write_messages(self, ws, session_data: SessionData, header_font, header_fill):
        """Write messages sheet."""
        headers = ["Message ID", "Parent ID", "Role", "Content", "Timestamp"]
        ws.append(headers)
        self._apply_header_style(ws, 1, header_font, header_fill)

        for msg in session_data.messages:
            ws.append([
                msg.message_id or "",
                msg.parent_id or "",
                msg.role,
                msg.content[:32767],  # Excel cell limit
                msg.timestamp.isoformat() if msg.timestamp else "",
            ])

        # Auto-width
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 25
